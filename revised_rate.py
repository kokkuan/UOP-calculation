"""
recalculate_uop.py

Recalculates Unit of Production (UOP) depreciation rates and amounts for
airport assets, using each airport's own passenger projections (2026-2069)
read from its REVISE RATE sheet.

Each airport has a different pax profile, so rates are derived per file.

Formula per asset (useful life = N years, starting 2026):
    total_pax       = sum of pax[2026 .. 2025+N]
    rate[year]      = pax[year] / total_pax   for year in [2026 .. 2025+N]
    rate[year]      = 0                        outside the asset's life
    dep_amount[year] = acquisition_value * rate[year]
"""

import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Files to process ──────────────────────────────────────────────────────────
MAIN_DIR     = Path(r"C:\DATA\Asset Calculation")
MA_SEPANG_DIR = MAIN_DIR / "MA SEPANG" / "CURRENT YEAR"
MASB_DIR     = MAIN_DIR / "MASB" / "CURRENT YEAR"
OUTPUT_DIR   = MAIN_DIR / "Output"

FILES = [
    MA_SEPANG_DIR / "KLIA 2026 - CUT OFF JUNE.xlsx",
    MA_SEPANG_DIR / "KLIA2 2026 - CUT OFF JUNE.xlsx",
    MASB_DIR / "006. KCH 2026 - CURRENT.xlsx",
    MASB_DIR / "011. LMN 2026 - CURRENT.xlsx",
    MASB_DIR / "014. MYY 2026 - CURRENT.xlsx",
    MASB_DIR / "018. SDK 2026 - CURRENT.xlsx",
    MASB_DIR / "019. SRWK 2026 - CURRENT.xlsx",
    MASB_DIR / "020. SZB 2026 - CURRENT.xlsx",
    MASB_DIR / "021. TGG 2026 - CURRENT.xlsx",
    MASB_DIR / "022. TWU 2026 - CURRENT.xlsx",
]

# ── Year range ────────────────────────────────────────────────────────────────
START_YEAR = 2026
END_YEAR   = 2069
YEARS      = list(range(START_YEAR, END_YEAR + 1))   # 2026 … 2069 inclusive
ALL_LIVES  = list(range(2, len(YEARS) + 1))           # 2 .. 44, full range, no skipping


def sheet_name_for(label: str) -> str:
    """Rate-matrix sheet name for an airport label, e.g. 'KLIA' -> 'KLIA'."""
    return label.replace(" ", "_")[:31]


def lookup_sheet_name_for(label: str) -> str:
    """RATE lookup sheet name for an airport label, e.g. 'KLIA' -> 'KLIA_RATE'."""
    return (label.replace(" ", "_") + "_RATE")[:31]

# ── UOP sheet: 0-based column indices ────────────────────────────────────────
#   (openpyxl cell column = 0-based index + 1)
COL_ASSET    = 1    # B  Asset number
COL_ACQUIS   = 5    # F  Acquisition value
COL_LIFE     = 13   # N  Useful life (years)

# Rate columns — note: 2026 is at index 14, then indices 15-16 hold the
# rounded rate and DepKy code (unchanged), so 2027 jumps to index 17
COL_RATE_2026 = 14  # O
COL_RATE_2027 = 17  # R  (indices 15=rounded, 16=DepKy are left untouched)
# 2028-2069 follow consecutively from index 18

COL_SUM_RATES  = 60  # BI  sum of all UOP rates (should equal 1.0)
COL_DIFF_RATES = 61  # BJ  difference = sum_rates - 1.0

# Depreciation amount columns (2026-2069 are consecutive starting at 62)
COL_DEP_START = 62   # BK  dep amount for 2026
# 2027 = 63, 2028 = 64 … 2069 = 105

COL_DEP_TOTAL = 106  # DK  total dep (should equal acquisition value)
COL_DEP_DIFF  = 107  # DL  difference = total_dep - acquisition_value

UOP_DATA_START_ROW = 6  # first asset data row in UOP sheet (1-based)

# ── Helpers ───────────────────────────────────────────────────────────────────

def rate_col(year: int) -> int:
    """Return 0-based column index for the UOP rate of a given year."""
    if year == START_YEAR:
        return COL_RATE_2026
    return COL_RATE_2027 + (year - (START_YEAR + 1))   # 2027 -> 17, 2028 -> 18 …


def dep_col(year: int) -> int:
    """Return 0-based column index for the dep amount of a given year."""
    return COL_DEP_START + (year - START_YEAR)          # 2026 -> 62, 2027 -> 63 …


# ── Core logic ────────────────────────────────────────────────────────────────

def read_pax(ws_revise_rate) -> dict:
    """
    Read annual passenger figures from REVISE RATE sheet.
    Returns {year: pax} for 2026-2069.
    """
    pax = {}
    for row in ws_revise_rate.iter_rows(min_row=7, max_row=50, values_only=True):
        year, passengers = row[3], row[4]
        if year and passengers:
            pax[int(year)] = float(passengers)
    return pax


def build_rate_matrix(pax: dict, useful_lives: set) -> dict:
    """
    Build UOP rate matrix keyed by useful life and year.
    rates[N][year] = pax[year] / sum(pax[2026 .. 2025+N])
    """
    rates = {}
    for N in useful_lives:
        life_years  = list(range(START_YEAR, START_YEAR + N))
        total_pax   = sum(pax[y] for y in life_years if y in pax)
        rates[N] = {}
        for y in YEARS:
            if y in life_years and total_pax > 0:
                rates[N][y] = pax[y] / total_pax
            else:
                rates[N][y] = 0.0
    return rates


def scan_uop_assets(ws_uop) -> list:
    """
    Read all asset rows from the UOP sheet.
    Returns list of (excel_row_number, useful_life, acquisition_value).
    """
    assets = []
    for excel_row, row in enumerate(
        ws_uop.iter_rows(min_row=UOP_DATA_START_ROW, values_only=True),
        start=UOP_DATA_START_ROW,
    ):
        asset_no = row[COL_ASSET]
        life     = row[COL_LIFE]
        acquis   = row[COL_ACQUIS]
        if asset_no is not None and life is not None and acquis is not None:
            assets.append((excel_row, int(life), float(acquis)))
    return assets


def read_rate_lookup(ws_rate) -> list:
    """
    Read the RATE sheet's Rate → DepKy lookup table as-is (no calculation).
    Returns list of (rate, depky) tuples, in original row order.
    """
    lookup = []
    for row in ws_rate.iter_rows(min_row=3, values_only=True):
        rate, depky = row[1], row[2]
        if rate is not None and depky is not None:
            lookup.append((rate, depky))
    return lookup


def write_results(ws_uop, assets: list, rates: dict):
    """
    Write recalculated rates and depreciation amounts to the UOP sheet.
    Leaves rounded-rate (col P) and DepKy (col Q) untouched.
    """
    for excel_row, life, acquis in assets:
        dep_total = 0.0

        for y in YEARS:
            r   = rates[life][y]
            dep = acquis * r

            # Write rate (openpyxl is 1-based, so col = 0-based index + 1)
            ws_uop.cell(row=excel_row, column=rate_col(y) + 1).value = r if r else None

            # Write dep amount
            ws_uop.cell(row=excel_row, column=dep_col(y) + 1).value = round(dep, 2) if dep else None

            dep_total += dep

        # Rate sum and difference
        sum_rates = sum(rates[life][y] for y in YEARS)
        ws_uop.cell(row=excel_row, column=COL_SUM_RATES  + 1).value = round(sum_rates, 10)
        ws_uop.cell(row=excel_row, column=COL_DIFF_RATES + 1).value = round(sum_rates - 1.0, 10)

        # Dep total and difference
        ws_uop.cell(row=excel_row, column=COL_DEP_TOTAL + 1).value = round(dep_total, 2)
        ws_uop.cell(row=excel_row, column=COL_DEP_DIFF  + 1).value = round(dep_total - acquis, 2)


# ── Rate summary Excel writer ─────────────────────────────────────────────────

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")   # mid blue
ZERO_FILL    = PatternFill("solid", fgColor="F2F2F2")   # light grey for zero cells
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
LABEL_FONT   = Font(bold=True, size=10)
THIN         = Side(style="thin", color="BFBFBF")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT_FORMAT   = "0.0000%"


def _cell(ws, row, col, value=None, fill=None, font=None, align="center", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:    c.fill      = fill
    if font:    c.font      = font
    if num_fmt: c.number_format = num_fmt
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = BORDER
    return c


def add_rate_sheet(wb_out, sheet_name: str, airport_label: str,
                   rates: dict, useful_lives: list, pax: dict):
    ws = wb_out.create_sheet(title=sheet_name)

    # ── Row 1: airport label spanning all columns ───────────────────────────
    total_cols = 1 + len(YEARS)   # "Useful Life" col + one col per year
    _cell(ws, 1, 1, airport_label, fill=HEADER_FILL, font=Font(bold=True, color="FFFFFF", size=11), align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # ── Row 2: pax projection ───────────────────────────────────────────────
    _cell(ws, 2, 1, "Pax projection", fill=SUBHDR_FILL, font=HEADER_FONT)
    for col_idx, y in enumerate(YEARS, start=2):
        _cell(ws, 2, col_idx, pax.get(y, 0), fill=SUBHDR_FILL, font=HEADER_FONT,
              num_fmt="#,##0")

    # ── Row 3: column headers (year) ────────────────────────────────────────
    _cell(ws, 3, 1, "Useful Life (yrs)", fill=SUBHDR_FILL, font=HEADER_FONT)
    for col_idx, y in enumerate(YEARS, start=2):
        _cell(ws, 3, col_idx, y, fill=SUBHDR_FILL, font=HEADER_FONT)

    # ── Rows 4+: one row per useful life ────────────────────────────────────
    for row_idx, N in enumerate(useful_lives, start=4):
        _cell(ws, row_idx, 1, N, font=LABEL_FONT, align="center")
        for col_idx, y in enumerate(YEARS, start=2):
            rate = rates[N][y]
            fill = ZERO_FILL if rate == 0 else None
            _cell(ws, row_idx, col_idx,
                  rate if rate else None,
                  fill=fill, num_fmt=PCT_FORMAT)

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 18
    for col_idx in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 16

    # Freeze panes: keep "Useful Life" column and header rows fixed
    ws.freeze_panes = "B4"


def add_rate_lookup_sheet(wb_out, sheet_name: str, airport_label: str, rate_lookup: list):
    """Write the RATE tab's Rate -> DepKy lookup table as-is, no calculation."""
    ws = wb_out.create_sheet(title=sheet_name)

    _cell(ws, 1, 1, airport_label, fill=HEADER_FILL, font=Font(bold=True, color="FFFFFF", size=11), align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    _cell(ws, 2, 1, "Rate", fill=SUBHDR_FILL, font=HEADER_FONT)
    _cell(ws, 2, 2, "DepKy", fill=SUBHDR_FILL, font=HEADER_FONT)

    for row_idx, (rate, depky) in enumerate(rate_lookup, start=3):
        _cell(ws, row_idx, 1, rate, num_fmt="0.0000")
        _cell(ws, row_idx, 2, depky)

    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 12
    ws.freeze_panes = "A3"


def write_rate_summary(all_airports: list, out_path: Path):
    """
    Write one Excel file with two sheets per airport:
      - the UOP rate matrix (rows = useful lives, columns = years 2026-2069, values = %)
      - the RATE tab's Rate -> DepKy lookup table, copied as-is (no calculation)
    all_airports: list of (label, rates_dict, useful_lives_list, pax_dict, rate_lookup)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    for label, rates, useful_lives, pax, rate_lookup in all_airports:
        add_rate_sheet(wb, sheet_name_for(label), label, rates, useful_lives, pax)
        add_rate_lookup_sheet(wb, lookup_sheet_name_for(label), airport_label=label, rate_lookup=rate_lookup)

    wb.save(out_path)
    print(f"\nRate summary saved: {out_path.name}")


# ── File processor ────────────────────────────────────────────────────────────

def process_file(filepath: Path) -> tuple:
    """Process one airport file. Returns (label, rates, useful_lives, pax, rate_lookup) for the summary."""
    print(f"\n{'='*60}")
    print(f"Processing: {filepath.name}")
    print(f"{'='*60}")

    # ── Step 1: Read pax projection (read-only for speed) ──────────────────
    wb_ro = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    pax   = read_pax(wb_ro["REVISE RATE"])
    wb_ro.close()

    print(f"\nStep 1 – Pax projection loaded:")
    print(f"  Years : {min(pax)} – {max(pax)}")
    print(f"  2026  : {pax[2026]:>16,.0f}")
    print(f"  Total : {sum(pax.values()):>16,.0f}")

    # ── Step 2: Scan UOP for unique useful lives ────────────────────────────
    wb_ro  = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    assets = scan_uop_assets(wb_ro["UOP"])
    wb_ro.close()

    # ── Step 2b: Read RATE tab lookup table as-is (no calculation) ──────────
    wb_ro = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    rate_lookup = read_rate_lookup(wb_ro["RATE"])
    wb_ro.close()
    print(f"\nStep 2b – RATE tab loaded: {len(rate_lookup)} rows (copied as-is)")

    lives_in_data = sorted({life for _, life, _ in assets})
    useful_lives  = ALL_LIVES   # full range 2 .. 44 (ends 2069), no skipping
    print(f"\nStep 2 – UOP assets found:")
    print(f"  Assets       : {len(assets)}")
    print(f"  Useful lives present in data : {lives_in_data}")
    print(f"  Rate matrix built for full range : {useful_lives[0]}-{useful_lives[-1]}")

    # ── Step 3: Build rate matrix ───────────────────────────────────────────
    rates = build_rate_matrix(pax, set(useful_lives))

    print(f"\nStep 3 – Rate matrix built (sample — first and last year of each life):")
    print(f"  {'Life':>4}  {'Start yr':>8}  {'End yr':>6}  {'Rate start':>12}  {'Rate end':>10}  {'Sum':>8}")
    print(f"  {'-'*4}  {'-'*8}  {'-'*6}  {'-'*12}  {'-'*10}  {'-'*8}")
    for N in useful_lives:
        start_rate = rates[N][START_YEAR]
        end_rate   = rates[N][START_YEAR + N - 1]
        total      = sum(rates[N][y] for y in YEARS)
        print(f"  {N:>4}  {START_YEAR:>8}  {START_YEAR+N-1:>6}  {start_rate:>12.6f}  {end_rate:>10.6f}  {total:>8.6f}")

    # derive airport label from filename stem (e.g. "KLIA" or "KLIA2")
    label_tokens = [t for t in filepath.stem.split(" ") if not re.match(r"^\d+\.$", t)]
    label = label_tokens[0]
    return (label, rates, useful_lives, pax, rate_lookup)


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    all_airports = []
    for f in FILES:
        if not f.exists():
            print(f"WARNING: File not found — {f}")
            continue
        result = process_file(f)
        all_airports.append(result)

    # Write combined UOP rate summary
    OUTPUT_DIR.mkdir(exist_ok=True)
    summary_path = OUTPUT_DIR / "Revised Rate.xlsx"
    write_rate_summary(all_airports, summary_path)

    print("\nAll files processed.")
