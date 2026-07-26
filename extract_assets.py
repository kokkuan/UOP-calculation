"""
extract_assets.py

Appends one "<label>_ASSETS" sheet per airport to Output\\PAX_Rate_Matrix_Test.xlsx
(must be built first by pax_rate_matrix.py), sourcing asset line items
from the consolidated "2.DATABASE" sheet (all airports in one SAP extract)
instead of per-airport UOP workbooks.

Each asset sheet has:
  - columns "Asset" .. "ORI. USEFUL LIFE" — copied as-is from 2.DATABASE,
    no calculation
  - "Remaining Useful Life" — REAL Excel formula: (year of Cap.Date + ORI.
    USEFUL LIFE) - 2026 + 1, i.e. years left of the asset's original life as
    of the 2026 cut-off, not counting the capitalization year itself as a
    life-year (life runs through cap_year + Use inclusive). If that's <= 0
    (life nominally already expired but the asset still carries real Book
    val. — true for ~11.5% of assets),
    it falls back to the full remaining concession window (MAX_LIFE, 44).
    Otherwise clamped to MAX_LIFE too, since the rate matrix only has rows
    1-44 (pax data only spans 2026-2069; any life >= 44 collapses to the
    same rate profile as life 44).
  - "New DepKy" — VLOOKUP of NEW RATE 2026 (rounded to 4dp) against the
    single global "DEPKEY_LOOKUP" sheet (copied from "3.DEPKEY FROM SAP"
    in the source workbook), giving the SAP depreciation key for the
    earliest year only
  - columns "NEW RATE" .. "Difference" — REAL Excel formulas (INDEX/MATCH,
    SUM), referencing the "<label>" rate-matrix sheet already in the same
    workbook, matched on Remaining Useful Life. Depreciation amounts are
    based on Book val. (net of Accum.dep. already recognized), not
    Acquis.val. — depreciating the full acquisition cost again from 2026
    would double-count whatever's already been expensed pre-2026.
  - a "TO NOTE" sheet listing every asset whose New DepKy could not be
    matched (its computed rate has no equal entry in 3.DEPKEY FROM SAP).

Run pax_rate_matrix.py first — this script only appends to its output.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from pax_rate_matrix import (
    SRC, OUTPUT_DIR, START_YEAR, YEARS, ALL_LIVES, MAX_LIFE,
    sheet_name_for, build_rate_matrix, load_source,
    read_pax_totals, read_mppa, apply_capping,
    read_masterfile, build_airport_key_map,
)

RATE_MATRIX_PATH = OUTPUT_DIR / "PAX_Rate_Matrix_Test.xlsx"

# ── Asset-info header labels, in 2.DATABASE column order ─────────────────────
ASSET_HEADERS = [
    "Asset", "SNo.", "Cap.Date", "Asset Description", "Acquis.val.", "Accum.dep.",
    "Book val.", "Crcy", "BusA", "APC", "Class", "DepKy", "ORI. USEFUL LIFE",
]
ASSET_INFO_COLS = len(ASSET_HEADERS)   # 13
COL_CAPDATE_IDX = 2    # index of "Cap.Date" within the asset-row tuple
COL_BOOKVAL_IDX = 6    # index of "Book val." within the asset-row tuple
COL_LIFE_IDX    = 12   # index of "ORI. USEFUL LIFE" within the asset-row tuple

# ── Output sheet column layout (1-based) ─────────────────────────────────────
# Block order: asset info | Remaining Useful Life | New DepKy | NEW RATE (all years) | Sum/Diff | Dep Amount (all years) | Total/Diff
N_YEARS         = len(YEARS)
REMLIFE_COL     = ASSET_INFO_COLS + 1
NEWDEPKY_COL    = REMLIFE_COL + 1
NEW_RATE_START  = NEWDEPKY_COL + 1
SUM_COL         = NEW_RATE_START + N_YEARS
DIFF_COL        = SUM_COL + 1
DEP_START       = DIFF_COL + 1
TOTAL_COL       = DEP_START + N_YEARS
DIFFTOTAL_COL   = TOTAL_COL + 1

CAPDATE_COL_LETTER = get_column_letter(COL_CAPDATE_IDX + 1)   # C
LIFE_COL_LETTER    = get_column_letter(COL_LIFE_IDX + 1)      # M
BOOKVAL_COL_LETTER = get_column_letter(COL_BOOKVAL_IDX + 1)   # G
REMLIFE_COL_LETTER = get_column_letter(REMLIFE_COL)           # N

# ── Rate-matrix sheet layout (built by pax_rate_matrix.py's add_rate_sheet) ──
RATE_MATRIX_LAST_COL   = get_column_letter(1 + N_YEARS)        # AS
RATE_MATRIX_DATA_ROW1  = 4
RATE_MATRIX_DATA_ROWN  = 3 + len(ALL_LIVES)                    # 47 (lives 1-44)

DEPKEY_LOOKUP_SHEET = "DEPKEY_LOOKUP"

DIVIDER_SHEET_NAME = "> Asset List"
DIVIDER_FILL       = PatternFill("solid", fgColor="ED7D31")   # orange, stands out from the blue rate sheets

SUMMARY_SHEET_NAME = "Summary"
TOTAL_FILL         = PatternFill("solid", fgColor="D9E1F2")   # light blue for the grand-total row

TO_NOTE_SHEET_NAME = "TO NOTE"
TO_NOTE_TAB_COLOR  = "C00000"                                  # dark red tab — flags items needing attention
TO_NOTE_HEADER_FILL = PatternFill("solid", fgColor="C00000")

# ── Styling (matches pax_rate_matrix.py) ─────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
LABEL_FONT  = Font(bold=True, size=10)
THIN        = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _cell(ws, row, col, value=None, fill=None, font=None, align="center", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:    c.fill      = fill
    if font:    c.font      = font
    if num_fmt: c.number_format = num_fmt
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = BORDER
    return c


def build_full_headers() -> list:
    headers = list(ASSET_HEADERS)
    headers += ["Remaining Useful Life"]
    headers += ["New DepKy"]
    headers += [f"NEW RATE {y}" for y in YEARS]
    headers += ["Sum UOP Rates", "Diff (Sum-1)"]
    headers += [f"Dep Amount {y}" for y in YEARS]
    headers += ["Total Dep", "Diff (Total-BookVal)"]
    return headers


def write_asset_row(ws, row_idx: int, asset_row: tuple, rate_sheet: str):
    # ── Asset info columns (raw values, as-is) ───────────────────────────────
    for col_idx, val in enumerate(asset_row, start=1):
        align = "left" if isinstance(val, str) else "center"
        _cell(ws, row_idx, col_idx, val, align=align)

    # ── Remaining Useful Life = (year of Cap.Date + ORI. USEFUL LIFE) - 2026 + 1 ──
    # Cap.Date is stored as text "DD.MM.YYYY", so pull the year via RIGHT(...,4).
    # The capitalization year itself doesn't count as a life-year (the asset's
    # life runs through cap_year + Use inclusive), so +1 vs. counting cap_year
    # as year 1. If the original life has nominally already expired (<=0) but
    # the asset still carries Book val., fall back to the remaining concession window.
    capdate_ref = f"${CAPDATE_COL_LETTER}{row_idx}"
    life_end_ref = f"(VALUE(RIGHT({capdate_ref},4))+${LIFE_COL_LETTER}{row_idx}-{START_YEAR}+1)"
    remlife_formula = f"=IF({life_end_ref}<=0,{MAX_LIFE},MIN({life_end_ref},{MAX_LIFE}))"
    _cell(ws, row_idx, REMLIFE_COL, remlife_formula)

    life_ref = f"${REMLIFE_COL_LETTER}{row_idx}"
    book_ref = f"${BOOKVAL_COL_LETTER}{row_idx}"

    # ── New DepKy: VLOOKUP of NEW RATE 2026 (rounded) against the global DepKy lookup ──
    new_rate_2026_ref = f"{get_column_letter(NEW_RATE_START)}{row_idx}"
    depkey_formula = (
        f"=IFERROR(VLOOKUP(ROUND({new_rate_2026_ref},4),"
        f"'{DEPKEY_LOOKUP_SHEET}'!$A:$B,2,FALSE),\"\")"
    )
    _cell(ws, row_idx, NEWDEPKY_COL, depkey_formula)

    # ── NEW RATE per year: INDEX/MATCH against the rate-matrix sheet ────────
    new_rate_refs = []
    for i, y in enumerate(YEARS):
        col = NEW_RATE_START + i
        formula = (
            f"=IFERROR(INDEX('{rate_sheet}'!$B${RATE_MATRIX_DATA_ROW1}:"
            f"${RATE_MATRIX_LAST_COL}${RATE_MATRIX_DATA_ROWN},"
            f"MATCH({life_ref},'{rate_sheet}'!$A${RATE_MATRIX_DATA_ROW1}:"
            f"$A${RATE_MATRIX_DATA_ROWN},0),"
            f"MATCH({y},'{rate_sheet}'!$B$3:${RATE_MATRIX_LAST_COL}$3,0)),0)"
        )
        _cell(ws, row_idx, col, formula, num_fmt="0.0000%")
        new_rate_refs.append(f"{get_column_letter(col)}{row_idx}")

    # ── Sum of rates / difference from 1.0 ───────────────────────────────────
    sum_range = f"{get_column_letter(NEW_RATE_START)}{row_idx}:{get_column_letter(NEW_RATE_START + N_YEARS - 1)}{row_idx}"
    _cell(ws, row_idx, SUM_COL, f"=SUM({sum_range})", num_fmt="0.0000%")
    _cell(ws, row_idx, DIFF_COL, f"={get_column_letter(SUM_COL)}{row_idx}-1", num_fmt="0.0000%")

    # ── Dep amount per year = Book val. (net of prior Accum.dep.) * rate ─────
    for i in range(N_YEARS):
        col = DEP_START + i
        formula = f"={book_ref}*{new_rate_refs[i]}"
        _cell(ws, row_idx, col, formula, num_fmt="#,##0.00")

    # ── Total dep / difference from Book val. ────────────────────────────────
    dep_range = f"{get_column_letter(DEP_START)}{row_idx}:{get_column_letter(DEP_START + N_YEARS - 1)}{row_idx}"
    _cell(ws, row_idx, TOTAL_COL, f"=SUM({dep_range})", num_fmt="#,##0.00")
    _cell(ws, row_idx, DIFFTOTAL_COL, f"={get_column_letter(TOTAL_COL)}{row_idx}-{book_ref}", num_fmt="#,##0.00")


def add_divider_sheet(wb_out, sheet_name: str):
    """A visual separator tab marking where the rate/lookup sheets end and the asset listings begin."""
    if sheet_name in wb_out.sheetnames:
        wb_out.remove(wb_out[sheet_name])
    ws = wb_out.create_sheet(title=sheet_name)

    _cell(ws, 2, 2, "Asset Listing", fill=DIVIDER_FILL, font=Font(bold=True, color="FFFFFF", size=18), align="left")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=6)
    _cell(ws, 4, 2, "Sheets to the right list every asset line item with recalculated UOP rates.",
          font=Font(italic=True, size=10), align="left")
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=6)

    for cell_row in (2, 4):
        for c in range(2, 7):
            ws.cell(row=cell_row, column=c).border = Border()

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "ED7D31"
    ws.column_dimensions[get_column_letter(1)].width = 3
    for col_idx in range(2, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.row_dimensions[2].height = 26


def add_depkey_lookup_sheet(wb_out, sheet_name: str, depkey_lookup: dict):
    """Write the global Rate -> DepKy lookup table (from '3.DEPKEY FROM SAP') as-is, no calculation."""
    if sheet_name in wb_out.sheetnames:
        wb_out.remove(wb_out[sheet_name])
    ws = wb_out.create_sheet(title=sheet_name)

    _cell(ws, 1, 1, "DepKy Lookup (from 3.DEPKEY FROM SAP)", fill=HEADER_FILL,
          font=Font(bold=True, color="FFFFFF", size=11), align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    _cell(ws, 2, 1, "Rate", fill=SUBHDR_FILL, font=HEADER_FONT)
    _cell(ws, 2, 2, "DepKy", fill=SUBHDR_FILL, font=HEADER_FONT)

    for row_idx, (rate, depky) in enumerate(sorted(depkey_lookup.items()), start=3):
        _cell(ws, row_idx, 1, rate, num_fmt="0.0000")
        _cell(ws, row_idx, 2, depky)

    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 12
    ws.freeze_panes = "A3"


def add_asset_sheet(wb_out, sheet_name: str, airport_label: str, headers: list, rows: list, rate_sheet: str):
    if sheet_name in wb_out.sheetnames:
        wb_out.remove(wb_out[sheet_name])
    ws = wb_out.create_sheet(title=sheet_name)

    total_cols = len(headers)
    _cell(ws, 1, 1, airport_label, fill=HEADER_FILL, font=Font(bold=True, color="FFFFFF", size=11), align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    for col_idx, h in enumerate(headers, start=1):
        _cell(ws, 2, col_idx, h, fill=SUBHDR_FILL, font=HEADER_FONT)

    for row_idx, asset_row in enumerate(rows, start=3):
        write_asset_row(ws, row_idx, asset_row, rate_sheet)

    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(11, len(h) + 2)

    ws.freeze_panes = "C3"


def add_summary_sheet(wb_out, sheet_name: str, airports: list):
    """
    Summary tab: rows = airport, columns = years 2026-2069, values = total
    depreciation for that airport/year, plus a per-airport row total and a
    grand-total row across all airports. Each cell is a real formula
    (full-column SUM against that airport's "<label>_ASSETS" sheet).
    airports: list of (label, assets_sheet_name)
    """
    if sheet_name in wb_out.sheetnames:
        wb_out.remove(wb_out[sheet_name])
    ws = wb_out.create_sheet(title=sheet_name, index=0)

    total_col   = 2 + N_YEARS   # one "Airport" col + N_YEARS year cols, then Total
    total_cols  = total_col

    _cell(ws, 1, 1, "Depreciation Summary by Airport (2026 - 2069)",
          fill=HEADER_FILL, font=Font(bold=True, color="FFFFFF", size=13), align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    _cell(ws, 2, 1, "Airport", fill=SUBHDR_FILL, font=HEADER_FONT)
    for col_idx, y in enumerate(YEARS, start=2):
        _cell(ws, 2, col_idx, y, fill=SUBHDR_FILL, font=HEADER_FONT)
    _cell(ws, 2, total_col, "Total", fill=SUBHDR_FILL, font=HEADER_FONT)

    first_data_row = 3
    for row_idx, (label, assets_sheet_name) in enumerate(airports, start=first_data_row):
        _cell(ws, row_idx, 1, label, font=LABEL_FONT, align="left")
        for i, y in enumerate(YEARS):
            col = 2 + i
            dep_col_letter = get_column_letter(DEP_START + i)
            formula = f"=SUM('{assets_sheet_name}'!{dep_col_letter}:{dep_col_letter})"
            _cell(ws, row_idx, col, formula, num_fmt="#,##0.00")
        row_range = f"{get_column_letter(2)}{row_idx}:{get_column_letter(1 + N_YEARS)}{row_idx}"
        _cell(ws, row_idx, total_col, f"=SUM({row_range})", num_fmt="#,##0.00", font=LABEL_FONT)

    total_row = first_data_row + len(airports)
    _cell(ws, total_row, 1, "Total", fill=TOTAL_FILL, font=LABEL_FONT, align="left")
    for i in range(N_YEARS):
        col = 2 + i
        col_range = f"{get_column_letter(col)}{first_data_row}:{get_column_letter(col)}{total_row - 1}"
        _cell(ws, total_row, col, f"=SUM({col_range})", fill=TOTAL_FILL, num_fmt="#,##0.00", font=LABEL_FONT)
    grand_range = f"{get_column_letter(2)}{total_row}:{get_column_letter(1 + N_YEARS)}{total_row}"
    _cell(ws, total_row, total_col, f"=SUM({grand_range})", fill=TOTAL_FILL, num_fmt="#,##0.00", font=LABEL_FONT)

    ws.column_dimensions[get_column_letter(1)].width = 14
    for col_idx in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "B3"


def add_to_note_sheet(wb_out, sheet_name: str, gaps: list):
    """
    Lists every asset whose New DepKy could not be matched (its computed
    NEW RATE 2026 has no equal entry in the '3.DEPKEY FROM SAP' lookup),
    grouped by airport. Tab is colored so it stands out as needing review.
    """
    if sheet_name in wb_out.sheetnames:
        wb_out.remove(wb_out[sheet_name])
    ws = wb_out.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = TO_NOTE_TAB_COLOR

    headers = ["Airport", "Asset", "SNo.", "Cap.Date", "Asset Description",
               "Acquis.val.", "BusA", "Eff. Life", "New Rate 2026", "Note"]
    _cell(ws, 1, 1, "Assets with no matching New DepKy", fill=TO_NOTE_HEADER_FILL,
          font=Font(bold=True, color="FFFFFF", size=12), align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for col_idx, h in enumerate(headers, start=1):
        _cell(ws, 2, col_idx, h, fill=TO_NOTE_HEADER_FILL, font=HEADER_FONT)

    for row_idx, g in enumerate(gaps, start=3):
        _cell(ws, row_idx, 1, g["Airport"], align="left")
        _cell(ws, row_idx, 2, g["Asset"])
        _cell(ws, row_idx, 3, g["SNo."])
        _cell(ws, row_idx, 4, g["Cap.Date"])
        _cell(ws, row_idx, 5, g["Description"], align="left")
        _cell(ws, row_idx, 6, round(float(g["Acquis.val."] or 0), 2), num_fmt="#,##0.00")
        _cell(ws, row_idx, 7, g["BusA"])
        _cell(ws, row_idx, 8, g["Eff. Life"])
        _cell(ws, row_idx, 9, round(g["New Rate 2026"], 6), num_fmt="0.0000%")
        _cell(ws, row_idx, 10, "No matching rate in 3.DEPKEY FROM SAP", align="left")

    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(h) + 2)
    ws.column_dimensions[get_column_letter(5)].width = 30
    ws.column_dimensions[get_column_letter(10)].width = 34
    ws.freeze_panes = "A3"


# ── Source reads (2.DATABASE / 3.DEPKEY FROM SAP) ────────────────────────────

def read_depkey_lookup(ws) -> dict:
    """Return {round(rate, 4): DepKy} from '3.DEPKEY FROM SAP' (first match wins on dup rates)."""
    lookup = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        depky, rate = row[1], row[3]
        if depky is None or rate is None:
            continue
        r = round(float(rate), 4)
        if r not in lookup:
            lookup[r] = depky
    return lookup


def read_database_assets(ws):
    """Yield asset dicts from '2.DATABASE' (header at row 4, data from row 5)."""
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        asset = row[2]
        if asset is None:
            continue
        yield {
            "Asset": asset,
            "SNo.": row[3],
            "Cap.Date": row[4],
            "Description": row[5],
            "Acquis.val.": row[6],
            "Accum.dep.": row[7],
            "Book val.": row[8],
            "Crcy": row[9],
            "BusA": int(row[10]),
            "APC FI": row[11],
            "Class": row[12],
            "DepKy": row[13],
            "Use": int(row[14]),
        }


def compute_remaining_life(cap_date: str, use: int, max_life: int = MAX_LIFE) -> int:
    """
    Mirrors the "Remaining Useful Life" Excel formula in write_asset_row:
    (year of Cap.Date + ORI. USEFUL LIFE) - START_YEAR + 1, clamped to max_life,
    falling back to max_life if the original life has nominally already
    expired (<=0) but the asset still carries Book val. The +1 excludes the
    capitalization year itself from the life count (life runs through
    cap_year + use inclusive).
    """
    cap_year = int(str(cap_date)[-4:])
    life_end = cap_year + use - START_YEAR + 1
    if life_end <= 0:
        return max_life
    return min(life_end, max_life)


def _load_join_inputs(src):
    """Shared read of 1. PAX BUDGET / 1.MPPA / MASTERFILE / 3.DEPKEY FROM SAP / 2.DATABASE."""
    wb = load_source(src)
    pax = read_pax_totals(wb["1. PAX BUDGET"])
    caps = read_mppa(wb["1.MPPA"])
    busa_to_airport = read_masterfile(wb["MASTERFILE"])
    depkey_lookup = read_depkey_lookup(wb["3.DEPKEY FROM SAP"])
    assets = list(read_database_assets(wb["2.DATABASE"]))
    wb.close()

    capped = apply_capping(pax, caps)
    airport_key_map = build_airport_key_map(capped, set(busa_to_airport.values()))
    return busa_to_airport, airport_key_map, capped, depkey_lookup, assets


def find_new_depky_gaps(src=SRC) -> list:
    """
    Recompute each asset's NEW RATE 2026 in Python (mirroring the Excel
    formula in write_asset_row) and return the ones with no matching DepKy
    in the '3.DEPKEY FROM SAP' lookup — i.e. what the "New DepKy" column
    would render as blank.
    """
    busa_to_airport, airport_key_map, capped, depkey_lookup, assets = _load_join_inputs(src)

    rate_matrix_cache = {}
    gaps = []
    for a in assets:
        airport_name = busa_to_airport[a["BusA"]]
        canonical = airport_key_map[airport_name]
        if canonical not in rate_matrix_cache:
            rate_matrix_cache[canonical] = build_rate_matrix(capped[canonical], set(range(1, MAX_LIFE + 1)))

        eff_life = compute_remaining_life(a["Cap.Date"], a["Use"])
        rate_2026 = rate_matrix_cache[canonical][eff_life][START_YEAR]
        new_depky = depkey_lookup.get(round(rate_2026, 4), "")

        if not new_depky:
            gaps.append({**a, "Airport": canonical, "Eff. Life": eff_life, "New Rate 2026": rate_2026})

    return gaps


# ── Asset grouping ────────────────────────────────────────────────────────────

def group_assets_by_airport(src=SRC) -> tuple:
    """
    Read every asset from '2.DATABASE', join BusA -> Airport via MASTERFILE,
    and group into {canonical_airport: [asset_row_tuple, ...]} where each
    asset_row_tuple matches ASSET_HEADERS order.

    Returns (grouped, depkey_lookup).
    """
    busa_to_airport, airport_key_map, _capped, depkey_lookup, assets = _load_join_inputs(src)

    grouped = {}
    for a in assets:
        airport_name = busa_to_airport[a["BusA"]]
        canonical = airport_key_map[airport_name]
        row = (
            a["Asset"], a["SNo."], a["Cap.Date"], a["Description"], a["Acquis.val."],
            a["Accum.dep."], a["Book val."], a["Crcy"], a["BusA"], a["APC FI"],
            a["Class"], a["DepKy"], a["Use"],
        )
        grouped.setdefault(canonical, []).append(row)

    return grouped, depkey_lookup


# ── Whole-workbook builder (reused by the CLI entry point and app.py) ────────

def add_asset_sheets(wb, src=SRC, progress_cb=None):
    """
    Append DepKy lookup, divider, per-airport asset sheets, Summary and
    TO NOTE sheets to an existing workbook (normally the one built by
    pax_rate_matrix.build_rate_matrix_workbook). Mutates `wb` in place.

    If progress_cb is provided, it will be called as progress_cb("assets", i, total, label)
    once per airport in sorted(grouped), where i is 1-based index and total = len(grouped).
    The callback fires for each airport regardless of whether its rate-matrix sheet was found
    (including the skip/WARNING path when rate_sheet is missing from the workbook).
    """
    grouped, depkey_lookup = group_assets_by_airport(src)

    # Clean up any existing summary/divider/asset/lookup sheets first, so
    # re-running always rebuilds them at the end, in order.
    for name in (SUMMARY_SHEET_NAME, TO_NOTE_SHEET_NAME, DIVIDER_SHEET_NAME, DEPKEY_LOOKUP_SHEET):
        if name in wb.sheetnames:
            wb.remove(wb[name])
    for label in grouped:
        assets_sheet_name = (label.replace(" ", "_") + "_ASSETS")[:31]
        if assets_sheet_name in wb.sheetnames:
            wb.remove(wb[assets_sheet_name])

    add_depkey_lookup_sheet(wb, DEPKEY_LOOKUP_SHEET, depkey_lookup)
    add_divider_sheet(wb, DIVIDER_SHEET_NAME)

    full_headers = build_full_headers()
    built_airports = []
    log = []
    airport_labels = sorted(grouped)
    total = len(airport_labels)
    for i, label in enumerate(airport_labels, start=1):
        rows = grouped[label]
        rate_sheet = sheet_name_for(label)
        if rate_sheet not in wb.sheetnames:
            log.append(f"WARNING: {label} — rate-matrix sheet '{rate_sheet}' not found, skipping ({len(rows)} assets)")
            if progress_cb is not None:
                progress_cb("assets", i, total, label)
            continue

        log.append(f"{label:20s}: {len(rows)} asset line items")
        assets_sheet_name = (label.replace(" ", "_") + "_ASSETS")[:31]
        add_asset_sheet(wb, assets_sheet_name, label, full_headers, rows, rate_sheet)
        built_airports.append((label, assets_sheet_name))

        if progress_cb is not None:
            progress_cb("assets", i, total, label)

    add_summary_sheet(wb, SUMMARY_SHEET_NAME, built_airports)

    gaps = find_new_depky_gaps(src)
    add_to_note_sheet(wb, TO_NOTE_SHEET_NAME, gaps)
    wb.move_sheet(TO_NOTE_SHEET_NAME, offset=2 - len(wb.sheetnames))  # move to index 1, right after Summary
    log.append(f"\nTO NOTE: {len(gaps)} asset(s) with no matching New DepKy")

    return log


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if not RATE_MATRIX_PATH.exists():
        raise SystemExit(f"{RATE_MATRIX_PATH} not found — run pax_rate_matrix.py first.")

    wb = openpyxl.load_workbook(RATE_MATRIX_PATH)
    log = add_asset_sheets(wb, SRC)
    for line in log:
        print(line)

    wb.save(RATE_MATRIX_PATH)
    print(f"\nAsset listing (with formulas) added to: {RATE_MATRIX_PATH.name}")
