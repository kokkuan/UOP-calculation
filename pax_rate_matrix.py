"""
pax_rate_matrix.py

Foundational module for the UOP depreciation pipeline. Builds one UOP
rate-matrix sheet per airport (useful life x year, 2026-2069) from the
consolidated workbook:
    WORKING FOR UOP CALCULATION MA S MASB 21072026.xlsx

Pax per airport comes from "1. PAX BUDGET" Total rows, MPPA-capped for 2026
only (from "1.MPPA"; airports with no MPPA row, e.g. Mukah, are left
uncapped). Airport identity is resolved via "MASTERFILE" (BusA -> Airport),
matched case-insensitively against the pax data — this also gives the
canonical airport list (all 28) used by extract_assets.py to join
"2.DATABASE" assets to their airport's rate matrix.

This is the first step of the pipeline — run this before extract_assets.py.
"""

import io
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MAIN_DIR   = Path(r"C:\DATA\Asset Calculation")
OUTPUT_DIR = MAIN_DIR / "Output"
SRC        = MAIN_DIR / "WORKING FOR UOP CALCULATION MA S MASB 21072026.xlsx"
OUT        = OUTPUT_DIR / "PAX_Rate_Matrix_Test.xlsx"

# ── Year range ────────────────────────────────────────────────────────────────
START_YEAR = 2026
END_YEAR   = 2069
YEARS      = list(range(START_YEAR, END_YEAR + 1))   # 2026 … 2069 inclusive
ALL_LIVES  = list(range(1, len(YEARS) + 1))           # 1 .. 44, full range, no skipping
MAX_LIFE   = len(YEARS)                               # 44 -- any remaining life >= this collapses to the same rate profile

# "1. PAX BUDGET": every section (International/ASEAN/NON ASEAN/Domestic/Total)
# shares the same column layout: col0=Airport, col1=<section label>, col2=2026 ... col45=2069
YEAR_COL_START = 2  # 0-based index of year 2026

SKIP_AIRPORTS = {"Total", "MASB"}  # aggregate rows, not real airports

CAPPED_FILL = PatternFill("solid", fgColor="FFC000")  # orange — pax was replaced by MPPA cap
PAX_ROW = 2
YEAR_COL_START_CELL = 2  # column of START_YEAR in the pax projection row (1-based)

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
SUBHDR_FILL  = PatternFill("solid", fgColor="2E75B6")   # mid blue
ZERO_FILL    = PatternFill("solid", fgColor="F2F2F2")   # light grey for zero cells
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
LABEL_FONT   = Font(bold=True, size=10)
THIN         = Side(style="thin", color="BFBFBF")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PCT_FORMAT   = "0.0000%"


def _cell(ws, row, col, value=None, fill=None, font=None, align="center", num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill:    c.fill      = fill
    if font:    c.font      = font
    if num_fmt: c.number_format = num_fmt
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = BORDER
    return c


def load_source(src) -> openpyxl.Workbook:
    """Open the source workbook, read-only. `src` may be a Path or raw bytes (e.g. an upload)."""
    if isinstance(src, (bytes, bytearray)):
        return openpyxl.load_workbook(io.BytesIO(src), data_only=True, read_only=True)
    return openpyxl.load_workbook(src, data_only=True, read_only=True)


def sheet_name_for(label: str) -> str:
    """Rate-matrix sheet name for an airport label, e.g. 'Terminal 1' -> 'Terminal_1'."""
    return label.replace(" ", "_")[:31]


# ── Core reads ────────────────────────────────────────────────────────────────

def read_pax_totals(ws) -> dict:
    """Read the 'Total' section rows (col1 == 'Total') -> {airport: {year: pax}}."""
    pax = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        airport, label = row[0], row[1]
        if airport is None or label != "Total":
            continue
        if airport in SKIP_AIRPORTS:
            continue
        pax[airport] = {
            y: row[YEAR_COL_START + (y - START_YEAR)] or 0.0
            for y in YEARS
        }
    return pax


def read_mppa(ws) -> dict:
    """Read {airport: cap} from '1.MPPA'."""
    caps = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        airport, cap = row[1], row[2]
        if airport is None or airport == "Airport" or cap is None:
            continue
        caps[airport] = float(cap)
    return caps


def apply_capping(pax: dict, caps: dict) -> dict:
    """Apply MPPA cap to year 2026 only. Returns {airport: {year: capped_pax}}."""
    capped = {}
    for airport, series in pax.items():
        new_series = dict(series)
        cap = caps.get(airport)
        if cap is not None and new_series[START_YEAR] > cap:
            new_series[START_YEAR] = cap
        capped[airport] = new_series
    return capped


def read_masterfile(ws) -> dict:
    """Return {BusA (int): Airport name} from MASTERFILE."""
    busa_to_airport = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        airport, busa = row[1], row[2]
        if airport is None or busa is None:
            continue
        busa_to_airport[int(busa)] = airport
    return busa_to_airport


def build_airport_key_map(pax: dict, masterfile_airports: set) -> dict:
    """Map MASTERFILE airport name -> canonical key used in the pax dict (case-insensitive)."""
    by_lower = {a.lower(): a for a in pax}
    key_map = {}
    for name in masterfile_airports:
        canonical = by_lower.get(name.lower())
        if canonical is None:
            raise KeyError(f"MASTERFILE airport {name!r} has no match in PAX BUDGET Total rows")
        key_map[name] = canonical
    return key_map


def build_rate_matrix(pax: dict, useful_lives: set) -> dict:
    """
    Build UOP rate matrix keyed by useful life and year.
    rates[N][year] = pax[year] / sum(pax[2026 .. 2025+N])
    """
    rates = {}
    for N in useful_lives:
        life_years  = list(range(START_YEAR, START_YEAR + N))
        total_pax   = sum(pax[y] for y in life_years if y in pax)
        rates[N] = {}
        for y in YEARS:
            if y in life_years and total_pax > 0:
                rates[N][y] = pax[y] / total_pax
            else:
                rates[N][y] = 0.0
    return rates


# ── Rate-matrix sheet writer ──────────────────────────────────────────────────

def add_rate_sheet(wb_out, sheet_name: str, airport_label: str,
                   rates: dict, useful_lives: list, pax: dict):
    if sheet_name in wb_out.sheetnames:
        wb_out.remove(wb_out[sheet_name])
    ws = wb_out.create_sheet(title=sheet_name)

    # ── Row 1: airport label spanning all columns ───────────────────────────
    total_cols = 1 + len(YEARS)   # "Useful Life" col + one col per year
    _cell(ws, 1, 1, airport_label, fill=HEADER_FILL, font=Font(bold=True, color="FFFFFF", size=11), align="left")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # ── Row 2: pax projection ───────────────────────────────────────────────
    _cell(ws, 2, 1, "Pax projection", fill=SUBHDR_FILL, font=HEADER_FONT)
    for col_idx, y in enumerate(YEARS, start=2):
        _cell(ws, 2, col_idx, pax.get(y, 0), fill=SUBHDR_FILL, font=HEADER_FONT,
              num_fmt="#,##0")

    # ── Row 3: column headers (year) ────────────────────────────────────────
    _cell(ws, 3, 1, "Useful Life (yrs)", fill=SUBHDR_FILL, font=HEADER_FONT)
    for col_idx, y in enumerate(YEARS, start=2):
        _cell(ws, 3, col_idx, y, fill=SUBHDR_FILL, font=HEADER_FONT)

    # ── Rows 4+: one row per useful life ────────────────────────────────────
    for row_idx, N in enumerate(useful_lives, start=4):
        _cell(ws, row_idx, 1, N, font=LABEL_FONT, align="center")
        for col_idx, y in enumerate(YEARS, start=2):
            rate = rates[N][y]
            fill = ZERO_FILL if rate == 0 else None
            _cell(ws, row_idx, col_idx,
                  rate if rate else None,
                  fill=fill, num_fmt=PCT_FORMAT)

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 18
    for col_idx in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 16

    # Freeze panes: keep "Useful Life" column and header rows fixed
    ws.freeze_panes = "B4"


# ── Whole-workbook builder (reused by the CLI entry point and app.py) ────────

def build_rate_matrix_workbook(src=SRC, progress_cb=None):
    """
    Read the source workbook and build a fresh workbook with one rate-matrix
    sheet per MASTERFILE airport (all 28). Returns (wb_out, airport_key_map).

    If progress_cb is provided, it will be called once per airport with the
    airport name as argument.
    """
    wb = load_source(src)
    pax = read_pax_totals(wb["1. PAX BUDGET"])
    caps = read_mppa(wb["1.MPPA"])
    busa_to_airport = read_masterfile(wb["MASTERFILE"])
    wb.close()

    capped = apply_capping(pax, caps)
    airport_key_map = build_airport_key_map(capped, set(busa_to_airport.values()))
    matched_airports = sorted(set(airport_key_map.values()))

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    log = []
    for airport in matched_airports:
        rates = build_rate_matrix(capped[airport], set(ALL_LIVES))
        sheet_name = sheet_name_for(airport)
        add_rate_sheet(wb_out, sheet_name, airport, rates, ALL_LIVES, capped[airport])

        was_capped = pax[airport][START_YEAR] != capped[airport][START_YEAR]
        if was_capped:
            wb_out[sheet_name].cell(row=PAX_ROW, column=YEAR_COL_START_CELL).fill = CAPPED_FILL
        log.append(f"Sheet added: {airport}{'  (2026 capped — highlighted)' if was_capped else ''}")

        if progress_cb is not None:
            progress_cb(airport)

    return wb_out, airport_key_map, log


def main():
    wb_out, _, log = build_rate_matrix_workbook(SRC)
    for line in log:
        print(line)

    OUTPUT_DIR.mkdir(exist_ok=True)
    wb_out.save(OUT)
    print(f"\nRate matrix workbook saved: {OUT}")


if __name__ == "__main__":
    main()
